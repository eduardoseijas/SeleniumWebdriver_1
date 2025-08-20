import time
#funciones excel
#from Funciones_Excel import *

#import unittest
from datetime import timezone
from os import times
from selenium import webdriver
from selenium.common import TimeoutException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.expected_conditions import alert_is_present
from selenium.webdriver.support.select import Select
#Son necesarias estas 2 librerias para el explicity_wait
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
#para que funcione el Select
from selenium.webdriver.support.ui import Select
#Esta es para el TimeoutException
from selenium.common.exceptions import TimeoutException
from AA_Libreria_Funciones_Globales.Libreria_Funciones import Libreria_Funciones_Globales
from AA_Libreria_Funciones_Globales.Page_loguinglobal import Page_Loguinglobal
from selenium.webdriver import ActionChains
import pytest #para importar la libreria pytest
from AA_Libreria_Funciones_Globales.Page_loguinglobal import Page_Loguinglobal
import openpyxl

class Funexcel():
    def __init__(self, driver):
        self.driver = driver

    def getRowCount(file,path,sheetName):
        Worbook=openpyxl.load_workbook(path)
        sheet= Worbook[sheetName]
        return (sheet.max_row)

    def getColumnCount(file,sheetName):
        Worbook= openpyxl.load_workbook(file)
        sheet=  Worbook[sheetName]
        return (sheet.max_column)

    def readData(file,path,sheetName,rownum,columnno):
        Worbook = openpyxl.load_workbook(path)
        sheet =  Worbook[sheetName]
        return sheet.cell(row=rownum, column=columnno).value

    def writeData(file,path,sheetName,rownum,columnno,data):
        Worbook = openpyxl.load_workbook(path)
        sheet =  Worbook[sheetName]
        sheet.cell(row=rownum, column=columnno).value = data
        Worbook.save(path)
